from flask import Flask, render_template, request, redirect, url_for, flash
import pandas as pd
from openpyxl import load_workbook
import os

from utils.WordConsolidator import consolidateWordOutputs
from utils.SampleProcessor import processSamples
from utils.SelectClient import SelectClient
from utils.PrepTemp import PrepTempFolder
from utils.TemplateWriter import fillTemplate
from params import debug

client_database = pd.read_excel("CellLineClients.xlsx")
client_list = client_database["Nickname"].tolist()

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'flask_app/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}
app.secret_key = 'some_secret_key'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    data = None
    selected_client = None
    po_number = "000000"
    can_query = False   # New flag to determine if the "Query Database" button should be displayed

    if request.method == 'POST':
        selected_client = request.form.get('client')
        po_number = request.form.get('po_number')

        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)

        file = request.files['file']

        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filename)

            app.config['FILENAME'] = file.filename

            # Read the data from the uploaded file
            wb = load_workbook(filename, read_only=True)
            ws = wb.active

            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)

            flash('File uploaded successfully!')

        if data and selected_client and po_number:
            can_query = True

    return render_template('upload_and_preview.html',
                           data=data, client_list=client_list,
                           selected_client=selected_client,
                           po_number=po_number, can_query=can_query)


@app.route('/preview/<filename>')
def preview_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    return render_template('preview.html', data=data)


@app.route('/query_database', methods=['POST'])
def query_database():
    print("Querying database...")
    selected_client = request.form.get('selected_client')
    po_number = request.form.get('po_number')

    # Load the data from the uploaded file into a dataframe
    file_path = app.config["UPLOAD_FOLDER"] + "/" + app.config["FILENAME"]

    print("File path: " + file_path)

    df = pd.read_excel(file_path)

    result_collection, sample_order = processSamples(df)

    # Display the results on the results page
    tables = []
    for each_sample in sample_order:
        for each in result_collection:
            results = each[0]
            sampleName = each[1]
            if sampleName == each_sample:
                print(type(results))
                tables.append(results)
                break

    # Set up the results page

    return render_template('results.html', tables=tables, selected_client=selected_client, po_number=po_number)


@app.route('/save_or_export', methods=['POST'])
def save_or_export():
    # You can receive the selected results from the form submission and then handle them appropriately.
    action = request.form.get('action')

    if action == "SAVE Project":
        # Handle saving the project
        # ...
        return "Project Saved!"
    elif action == "Export Results":
        # Handle exporting the results
        # ...
        return "Results Exported!"


if __name__ == '__main__':
    app.run(debug=True)
